#!/usr/bin/env python3
"""
certmon.py  –  SSL Certificate Monitor  v1
===========================================
Reads a list of URLs / hostnames from a TXT file (e.g. discovered_urls.txt
produced by domain_discovery.py), checks each one's SSL certificate and TLS
configuration, and writes the results to both CSV and XLSX.

Checks performed
----------------
Certificate:
  • Subject CN / O
  • Issuer CN / O
  • Not-Before / Not-After (validity window)
  • Days until expiry  (⚠ warning if < WARNING_DAYS, 🔴 critical if expired)
  • Serial number
  • Signature algorithm  (flags SHA-1, MD5 as weak)
  • Public key type + size  (flags RSA < 2048, EC < 224)
  • Subject Alternative Names (DNS + IP)
  • Hostname match
  • Wildcard certificate
  • Self-signed detection
  • Chain trust (via requests / system CA store)
  • OCSP URL presence
  • AIA Issuer URL presence
  • Basic Constraints CA flag

TLS protocol:
  • Negotiated protocol (TLS 1.0 / 1.1 / 1.2 / 1.3)
  • Deprecated TLS 1.0 / 1.1 accepted  (misconfiguration)
  • Selected cipher suite

HTTP:
  • HTTP status code
  • HTTP → HTTPS redirect check
  • HSTS header presence (Strict-Transport-Security)

Output columns (CSV + XLSX)
---------------------------
  url, host, port, scan_status, http_status, hsts, http_to_https_redirect,
  cert_subject_cn, cert_subject_o, cert_issuer_cn, cert_issuer_o,
  serial_number, sig_algorithm, public_key_type, public_key_bits,
  not_before, not_after, days_until_expiry, expiry_status,
  san_dns, san_ip, hostname_match, wildcard, self_signed, chain_trusted,
  ocsp_url, aia_issuer, is_ca,
  tls_negotiated, tls_supported, tls_deprecated_enabled,
  tls_minimum, tls_maximum, cipher_suite,
  deprecated_tls10, deprecated_tls11,
  issues, weaknesses,
  scanned_at

Usage
-----
  python3 certmon.py -i discovered_urls.txt
  python3 certmon.py -i discovered_urls.txt -o report.xlsx --threads 20 --warn-days 60
  python3 certmon.py -i discovered_urls.txt --csv-only
"""

import argparse
import concurrent.futures as cf
import csv
import datetime as dt
import json
import logging
import socket
import ssl
import sys
from dataclasses import dataclass, asdict, fields
from pathlib import Path
from typing import List, Optional, Tuple
from urllib.parse import urlparse

import requests
import urllib3
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from cryptography.x509.oid import ExtensionOID, NameOID
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────
LOG_FMT = "%(asctime)s [%(levelname)s] %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FMT)
log = logging.getLogger("certmon")

# ──────────────────────────────────────────────────────────────────────────────
# Defaults
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT_WARNING_DAYS = 30
DEFAULT_CRITICAL_DAYS = 0      # expired
DEFAULT_THREADS = 10   # keep lower to avoid self-induced timeouts on large scans
DEFAULT_TIMEOUT = 15
DEFAULT_RETRIES = 2       # retry on timeout before marking unreachable
# All HTTPS ports tried during SSL connect (in order).
# Plain HTTP counterparts are derived automatically: 443→80, 8443→8080, etc.
HTTPS_PORTS: List[int] = [443, 8443, 4443, 10443]
HTTP_PORTS:  dict      = {443: 80, 8443: 8080, 4443: 4080, 10443: 10080}
# HTTP ports probed for redirect check (if HTTPS port not in mapping, port-1 is tried)
HTTP_REDIRECT_PORTS: List[int] = [80, 8080, 8000, 8008]
USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) CertMon/1.0"

WEAK_HASH_ALGOS  = {"sha1", "md5", "sha1withrsa", "md5withrsa"}
WEAK_RSA_BITS    = 2048
WEAK_EC_BITS     = 224
# All TLS versions probed — used to build the full supported-versions matrix
ALL_TLS_VERSIONS = [
    ("TLS 1.0", ssl.TLSVersion.TLSv1),
    ("TLS 1.1", ssl.TLSVersion.TLSv1_1),
    ("TLS 1.2", ssl.TLSVersion.TLSv1_2),
    ("TLS 1.3", ssl.TLSVersion.TLSv1_3),
]
DEPRECATED_TLS = {"TLS 1.0", "TLS 1.1"}


# ──────────────────────────────────────────────────────────────────────────────
# Result dataclass  (one row per URL)
# ──────────────────────────────────────────────────────────────────────────────
@dataclass
class CertRow:
    url:                    str  = ""
    host:                   str  = ""
    port:                   int  = 443
    scan_status:            str  = ""   # ok | no_ssl | unreachable | error
    http_status:            str  = ""
    hsts:                   str  = ""   # yes | no | n/a
    http_to_https_redirect: str  = ""   # yes | no | n/a

    cert_subject_cn:        str  = ""
    cert_subject_o:         str  = ""
    cert_issuer_cn:         str  = ""
    cert_issuer_o:          str  = ""
    serial_number:          str  = ""
    sig_algorithm:          str  = ""
    public_key_type:        str  = ""
    public_key_bits:        str  = ""

    not_before:             str  = ""
    not_after:              str  = ""
    days_until_expiry:      str  = ""
    expiry_status:          str  = ""   # OK | WARNING | EXPIRED | NOT_YET_VALID

    san_dns:                str  = ""
    san_ip:                 str  = ""
    hostname_match:         str  = ""   # yes | no | n/a
    wildcard:               str  = ""   # yes | no
    self_signed:            str  = ""   # yes | no
    chain_trusted:          str  = ""   # yes | no | unknown

    ocsp_url:               str  = ""
    aia_issuer:             str  = ""
    is_ca:                  str  = ""   # yes | no

    tls_negotiated:         str  = ""   # exact version negotiated for THIS connection
    tls_supported:          str  = ""   # all versions accepted, e.g. "TLS 1.2; TLS 1.3"
    tls_deprecated_enabled: str  = ""   # deprecated versions accepted, e.g. "TLS 1.0; TLS 1.1" or ""
    tls_minimum:            str  = ""   # lowest accepted version
    tls_maximum:            str  = ""   # highest accepted version
    cipher_suite:           str  = ""
    deprecated_tls10:       str  = ""   # yes | no  (kept for backward compat / filters)
    deprecated_tls11:       str  = ""   # yes | no

    issues:                 str  = ""   # semicolon-separated critical issues
    weaknesses:             str  = ""   # semicolon-separated warnings
    scanned_at:             str  = ""
    redirect_chain:         str  = ""   # recorded hop-by-hop if redirect followed
    final_https_url:        str  = ""   # URL after all redirects resolved
    error_detail:           str  = ""


# ──────────────────────────────────────────────────────────────────────────────
# Input parsing
# ──────────────────────────────────────────────────────────────────────────────
def load_urls(path: Path) -> List[Tuple[str, str, int]]:
    """
    Returns list of (original_url, host, port).
    Accepts bare hostnames, https://host/path, http://host lines.
    Skips blank lines and # comments.
    """
    entries = []
    seen = set()
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "://" not in line:
            line = "https://" + line
        # Safety net: upgrade http:// to https:// — certmon only scans TLS
        if line.startswith("http://"):
            p0 = urlparse(line)
            netloc = p0.hostname or ""
            if p0.port and p0.port != 80:
                netloc = f"{netloc}:{p0.port}"
            line = urlunparse(("https", netloc, p0.path or "/", p0.params, p0.query, ""))
            log.debug("load_urls: upgraded http→https: %s", line)
        p = urlparse(line)
        host = (p.hostname or "").lower()
        # Use explicit port if given; otherwise 443 for https/bare
        port = p.port or 443
        if not host:
            continue
        key = (host, port)
        if key in seen:
            continue
        seen.add(key)
        entries.append((line, host, port))
    return entries


# ──────────────────────────────────────────────────────────────────────────────
# Certificate helpers
# ──────────────────────────────────────────────────────────────────────────────
def _attr(obj, oid) -> str:
    try:
        return obj.get_attributes_for_oid(oid)[0].value
    except Exception:
        return ""

def _san_dns(cert) -> List[str]:
    try:
        ext = cert.extensions.get_extension_for_oid(ExtensionOID.SUBJECT_ALTERNATIVE_NAME).value
        return list(ext.get_values_for_type(x509.DNSName))
    except Exception:
        return []

def _san_ip(cert) -> List[str]:
    try:
        ext = cert.extensions.get_extension_for_oid(ExtensionOID.SUBJECT_ALTERNATIVE_NAME).value
        return [str(x) for x in ext.get_values_for_type(x509.IPAddress)]
    except Exception:
        return []

def _aia_ocsp(cert) -> List[str]:
    try:
        ext = cert.extensions.get_extension_for_oid(ExtensionOID.AUTHORITY_INFORMATION_ACCESS).value
        return [d.access_location.value for d in ext
                if d.access_method.dotted_string == "1.3.6.1.5.5.7.48.1"]
    except Exception:
        return []

def _aia_issuers(cert) -> List[str]:
    try:
        ext = cert.extensions.get_extension_for_oid(ExtensionOID.AUTHORITY_INFORMATION_ACCESS).value
        return [d.access_location.value for d in ext
                if d.access_method.dotted_string == "1.3.6.1.5.5.7.48.2"]
    except Exception:
        return []

def _is_ca(cert) -> bool:
    try:
        return cert.extensions.get_extension_for_oid(
            ExtensionOID.BASIC_CONSTRAINTS).value.ca
    except Exception:
        return False

def _hostname_match(cert, host: str) -> bool:
    san = _san_dns(cert)
    try:
        ssl.match_hostname(
            {"subjectAltName": [("DNS", d) for d in san]}, host
        )
        return True
    except Exception:
        try:
            cn = _attr(cert.subject, NameOID.COMMON_NAME)
            ssl.match_hostname({"subject": ((("commonName", cn),),)}, host)
            return True
        except Exception:
            return False

def _is_self_signed(cert) -> bool:
    try:
        return cert.issuer == cert.subject
    except Exception:
        return False

def _sig_alg(cert) -> str:
    try:
        return cert.signature_hash_algorithm.name
    except Exception:
        return ""

def _pk_info(cert) -> Tuple[str, int]:
    try:
        pk = cert.public_key()
        name = type(pk).__name__
        size = getattr(pk, "key_size", 0)
        return name, size
    except Exception:
        return "", 0


# ──────────────────────────────────────────────────────────────────────────────
# TLS protocol probe
# ──────────────────────────────────────────────────────────────────────────────
def _supports_proto(host: str, port: int, proto: ssl.TLSVersion, timeout: int) -> bool:
    try:
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        ctx.minimum_version = proto
        ctx.maximum_version = proto
        with socket.create_connection((host, port), timeout=timeout) as sock:
            with ctx.wrap_socket(sock, server_hostname=host):
                return True
    except Exception:
        return False


# ──────────────────────────────────────────────────────────────────────────────
# HTTP checks
# ──────────────────────────────────────────────────────────────────────────────
def _https_url(host: str, port: int) -> str:
    return f"https://{host}/" if port == 443 else f"https://{host}:{port}/"


def _http_url(host: str, http_port: int) -> str:
    return f"http://{host}/" if http_port == 80 else f"http://{host}:{http_port}/"


def _probe_https(session: requests.Session, host: str, port: int, timeout: int) -> Tuple[str, str, str]:
    """
    Probe the HTTPS endpoint.
    Returns (http_status, hsts, final_url_after_redirects).
    Follows all redirects and records the final URL so we can detect
    cross-port / cross-scheme redirects.
    """
    url = _https_url(host, port)
    hsts = "no"
    http_status = ""
    final_url = url
    try:
        r = session.get(url, timeout=timeout, verify=True, allow_redirects=True)
        http_status = str(r.status_code)
        final_url   = r.url
        if "strict-transport-security" in r.headers:
            hsts = "yes"
        return http_status, hsts, final_url
    except requests.exceptions.SSLError:
        # Retry without verification to still get the HTTP status
        try:
            r = session.get(url, timeout=timeout, verify=False, allow_redirects=True)
            http_status = str(r.status_code)
            final_url   = r.url
            if "strict-transport-security" in r.headers:
                hsts = "yes"
        except Exception:
            http_status = "SSL_ERROR"
    except requests.exceptions.ConnectionError:
        http_status = "CONN_ERROR"
    except requests.exceptions.Timeout:
        http_status = "TIMEOUT"
    except Exception as exc:
        http_status = f"ERROR:{type(exc).__name__}"
    return http_status, hsts, final_url


def _probe_http_redirect(session: requests.Session, host: str,
                         https_port: int, timeout: int) -> str:
    """
    Try every candidate HTTP port to find a redirect to HTTPS.
    Returns "yes:<http_port>→<https_port>" | "no" | "n/a".

    Logic:
      1. Prefer the canonical HTTP counterpart of the HTTPS port
         (e.g. 8443 → 8080) if it is reachable.
      2. Fall back through HTTP_REDIRECT_PORTS.
      3. Accept both single-hop redirects (30x Location: https://…)
         and meta-refresh / JS redirects are NOT chased (HTTP layer only).
      4. Also accept redirect chains: http→http→https counts as "yes".
    """
    # Build ordered candidate list: canonical counterpart first
    canonical = HTTP_PORTS.get(https_port)
    candidates = []
    if canonical:
        candidates.append(canonical)
    for p in HTTP_REDIRECT_PORTS:
        if p not in candidates:
            candidates.append(p)

    for http_port in candidates:
        url = _http_url(host, http_port)
        try:
            # Do NOT follow redirects automatically — inspect each hop
            resp = session.get(url, timeout=timeout, verify=False,
                               allow_redirects=False)
            location = resp.headers.get("location", "")
            code     = resp.status_code

            # Direct single-hop redirect to HTTPS
            if code in (301, 302, 303, 307, 308) and location.startswith("https://"):
                return f"yes:{http_port}→https"

            # Follow redirect chain (up to 5 hops) to see if it ends at HTTPS
            if code in (301, 302, 303, 307, 308) and location:
                current = requests.compat.urljoin(url, location)
                for _ in range(5):
                    if current.startswith("https://"):
                        return f"yes:{http_port}→https(chain)"
                    try:
                        r2 = session.get(current, timeout=timeout, verify=False,
                                         allow_redirects=False)
                        loc2 = r2.headers.get("location", "")
                        if r2.status_code in (301, 302, 303, 307, 308) and loc2:
                            current = requests.compat.urljoin(current, loc2)
                        else:
                            break
                    except Exception:
                        break

            # Port is reachable but serves HTTP without redirect
            return "no"

        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout):
            continue   # port not open — try next
        except Exception:
            continue

    return "n/a"   # no HTTP port responded at all


def _http_checks(host: str, port: int, timeout: int) -> Tuple[str, str, str]:
    """
    Returns (http_status, hsts, http_to_https_redirect).
    http_to_https_redirect values:
      "yes:<src_port>→https"        — clean single-hop redirect
      "yes:<src_port>→https(chain)" — multi-hop chain ending at HTTPS
      "no"                          — HTTP port exists but no redirect
      "n/a"                         — no HTTP port found / responded
    """
    session = requests.Session()
    session.headers["User-Agent"] = USER_AGENT

    http_status, hsts, _ = _probe_https(session, host, port, timeout)
    http_to_https = _probe_http_redirect(session, host, port, timeout)
    return http_status, hsts, http_to_https


# ──────────────────────────────────────────────────────────────────────────────
# Chain trust
# ──────────────────────────────────────────────────────────────────────────────
def _chain_trusted(host: str, port: int, timeout: int) -> str:
    try:
        ctx = ssl.create_default_context()
        with socket.create_connection((host, port), timeout=timeout) as sock:
            sock.settimeout(timeout)
            with ctx.wrap_socket(sock, server_hostname=host):
                return "yes"
    except ssl.SSLCertVerificationError:
        return "no"
    except Exception:
        return "unknown"


# ──────────────────────────────────────────────────────────────────────────────
# Main scan function
# ──────────────────────────────────────────────────────────────────────────────
def scan_host(url: str, host: str, port: int, warn_days: int, timeout: int) -> CertRow:
    now_utc  = dt.datetime.now(dt.timezone.utc)
    scanned  = now_utc.isoformat()
    row      = CertRow(url=url, host=host, port=port, scanned_at=scanned)
    issues: List[str]    = []
    weaknesses: List[str] = []

    # ── Fetch raw certificate (no chain validation) ───────────────────────────
    # Tries FALLBACK_PORTS in order, retries up to DEFAULT_RETRIES on timeout.
    # Distinguishes: ssl_error | connection_refused | timeout→retry | unreachable
    der = cipher = version = None
    last_error = ""
    actual_port = port
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode    = ssl.CERT_NONE

    ports_to_try = [port] + [p for p in HTTPS_PORTS if p != port]
    connected = False
    for try_port in ports_to_try:
        for attempt in range(1, DEFAULT_RETRIES + 2):   # +2 → retries+1 attempts
            try:
                with socket.create_connection((host, try_port), timeout=timeout) as sock:
                    sock.settimeout(timeout)             # also covers SSL handshake
                    with ctx.wrap_socket(sock, server_hostname=host) as ssock:
                        der     = ssock.getpeercert(binary_form=True)
                        cipher  = ssock.cipher()
                        version = ssock.version()
                        actual_port = try_port
                        connected = True
                break   # success
            except ssl.SSLError as e:
                row.scan_status  = "no_ssl"
                row.error_detail = f"port {try_port}: {e}"
                row.issues = "no_ssl_handshake"
                return row
            except (socket.timeout, TimeoutError) as e:
                last_error = f"port {try_port} attempt {attempt}: timeout"
                log.debug("Timeout %s:%s attempt %d/%d", host, try_port, attempt, DEFAULT_RETRIES + 1)
                continue   # retry same port
            except ConnectionRefusedError as e:
                last_error = f"port {try_port}: connection refused"
                break      # no point retrying refused — try next port
            except OSError as e:
                last_error = f"port {try_port}: {e}"
                break      # DNS failure or similar — no point retrying
            except Exception as e:
                last_error = f"port {try_port}: {e}"
                break
        if connected:
            break

    if not connected:
        row.scan_status  = "unreachable"
        row.error_detail = last_error
        row.issues = ""
        return row

    row.port = actual_port

    row.scan_status   = "ok"
    row.tls_negotiated = version or ""   # set here; full matrix built later
    row.cipher_suite  = cipher[0] if cipher else ""

    # ── Parse certificate ─────────────────────────────────────────────────────
    cert = x509.load_der_x509_certificate(der, default_backend())

    row.cert_subject_cn = _attr(cert.subject, NameOID.COMMON_NAME)
    row.cert_subject_o  = _attr(cert.subject, NameOID.ORGANIZATION_NAME)
    row.cert_issuer_cn  = _attr(cert.issuer,  NameOID.COMMON_NAME)
    row.cert_issuer_o   = _attr(cert.issuer,  NameOID.ORGANIZATION_NAME)
    row.serial_number   = hex(cert.serial_number)

    sig = _sig_alg(cert)
    row.sig_algorithm   = sig
    if sig.lower() in WEAK_HASH_ALGOS:
        weaknesses.append(f"weak_signature_algorithm:{sig}")

    pk_type, pk_bits    = _pk_info(cert)
    row.public_key_type = pk_type
    row.public_key_bits = str(pk_bits)
    if "rsa" in pk_type.lower() and pk_bits and pk_bits < WEAK_RSA_BITS:
        weaknesses.append(f"weak_rsa_key:{pk_bits}bits")
    elif "ec" in pk_type.lower() and pk_bits and pk_bits < WEAK_EC_BITS:
        weaknesses.append(f"weak_ec_key:{pk_bits}bits")

    # ── Validity dates ────────────────────────────────────────────────────────
    nb = cert.not_valid_before_utc
    na = cert.not_valid_after_utc
    row.not_before = nb.strftime("%Y-%m-%d")
    row.not_after  = na.strftime("%Y-%m-%d")
    days_left      = (na - now_utc).days
    row.days_until_expiry = str(days_left)

    if na <= now_utc:
        row.expiry_status = "EXPIRED"
        issues.append(f"certificate_expired_since_{(-days_left)}d")
    elif days_left <= warn_days:
        row.expiry_status = "WARNING"
        issues.append(f"expiring_in_{days_left}d")
    else:
        row.expiry_status = "OK"

    if nb > now_utc:
        row.expiry_status = "NOT_YET_VALID"
        issues.append("certificate_not_yet_valid")

    # ── SAN ───────────────────────────────────────────────────────────────────
    san_dns = _san_dns(cert)
    san_ip  = _san_ip(cert)
    row.san_dns = "; ".join(san_dns)
    row.san_ip  = "; ".join(san_ip)

    if not san_dns:
        weaknesses.append("missing_san_dns_extension")

    # ── Hostname match ────────────────────────────────────────────────────────
    row.hostname_match = "yes" if _hostname_match(cert, host) else "no"
    if row.hostname_match == "no":
        issues.append("hostname_mismatch")

    # ── Wildcard ──────────────────────────────────────────────────────────────
    row.wildcard = "yes" if any(d.startswith("*.") for d in san_dns) else "no"
    if row.wildcard == "yes":
        weaknesses.append("wildcard_certificate")

    # ── Self-signed ───────────────────────────────────────────────────────────
    row.self_signed = "yes" if _is_self_signed(cert) else "no"
    if row.self_signed == "yes":
        issues.append("self_signed_certificate")

    # ── Chain trust ───────────────────────────────────────────────────────────
    row.chain_trusted = _chain_trusted(host, port, timeout)
    if row.chain_trusted == "no":
        issues.append("untrusted_certificate_chain")

    # ── OCSP / AIA ────────────────────────────────────────────────────────────
    ocsp    = _aia_ocsp(cert)
    issuers = _aia_issuers(cert)
    row.ocsp_url   = "; ".join(ocsp)
    row.aia_issuer = "; ".join(issuers)

    if not ocsp:
        weaknesses.append("missing_ocsp_url")
    if not issuers:
        weaknesses.append("missing_aia_issuer")

    # ── CA flag ───────────────────────────────────────────────────────────────
    row.is_ca = "yes" if _is_ca(cert) else "no"

    # ── Full TLS version matrix ───────────────────────────────────────────────
    # Probe every known TLS version to build a complete picture.
    supported_versions: List[str] = []
    for ver_label, ver_enum in ALL_TLS_VERSIONS:
        try:
            if _supports_proto(host, actual_port, ver_enum, timeout):
                supported_versions.append(ver_label)
        except Exception:
            pass

    deprecated_found = [v for v in supported_versions if v in DEPRECATED_TLS]
    row.tls_negotiated         = version or ""     # version from the initial handshake
    row.tls_supported          = "; ".join(supported_versions) if supported_versions else "unknown"
    row.tls_deprecated_enabled = "; ".join(deprecated_found)
    row.tls_minimum            = supported_versions[0]  if supported_versions else ""
    row.tls_maximum            = supported_versions[-1] if supported_versions else ""
    row.deprecated_tls10       = "yes" if "TLS 1.0" in deprecated_found else "no"
    row.deprecated_tls11       = "yes" if "TLS 1.1" in deprecated_found else "no"
    if "TLS 1.0" in deprecated_found:
        weaknesses.append("deprecated_TLS_1.0_accepted")
    if "TLS 1.1" in deprecated_found:
        weaknesses.append("deprecated_TLS_1.1_accepted")
    if supported_versions and row.tls_maximum not in ("TLS 1.2", "TLS 1.3"):
        issues.append(f"no_modern_tls_supported")
    if "TLS 1.3" not in supported_versions:
        weaknesses.append("TLS_1.3_not_supported")

    # ── HTTP layer checks ─────────────────────────────────────────────────────
    http_status, hsts, http_to_https = _http_checks(host, actual_port, timeout)
    row.http_status            = http_status
    row.hsts                   = hsts
    row.http_to_https_redirect = http_to_https

    # Record redirect chain and final URL for transparency
    _session = requests.Session()
    _session.headers["User-Agent"] = USER_AGENT
    try:
        _r = _session.get(_https_url(host, actual_port), timeout=timeout,
                          verify=False, allow_redirects=True)
        row.final_https_url = _r.url
        row.redirect_chain  = " → ".join(
            str(h.url) for h in _r.history
        ) if _r.history else ""
    except Exception:
        pass

    if hsts == "no":
        weaknesses.append("missing_hsts_header")
    if http_to_https == "no":
        weaknesses.append("no_http_to_https_redirect")

    row.issues     = "; ".join(sorted(set(issues)))
    row.weaknesses = "; ".join(sorted(set(weaknesses)))
    return row


# ──────────────────────────────────────────────────────────────────────────────
# CSV output
# ──────────────────────────────────────────────────────────────────────────────
def write_csv(rows: List[CertRow], path: Path) -> None:
    col_names = [f.name for f in fields(CertRow)]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=col_names)
        w.writeheader()
        for r in rows:
            w.writerow(asdict(r))
    log.info("CSV saved → %s", path)


# ──────────────────────────────────────────────────────────────────────────────
# XLSX output
# ──────────────────────────────────────────────────────────────────────────────
# Colour palette (dark-navy header, traffic-light for expiry/issues)
HDR_BG   = "1F4E78"
HDR_FG   = "FFFFFF"
RED_BG   = "FFCCCC"
AMBER_BG = "FFE5CC"
GREEN_BG = "CCFFCC"
GREY_BG  = "F2F2F2"

def _col_width(rows: List[CertRow], col: str, header: str) -> float:
    values = [str(getattr(r, col) or "") for r in rows]
    max_len = max((len(v) for v in values), default=0)
    return min(max(max_len, len(header)) + 2, 55)


def write_xlsx(rows: List[CertRow], path: Path, warn_days: int) -> None:
    wb = Workbook()

    # ── Sheet 1: Full results ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Certificate Scan"

    col_names = [f.name for f in fields(CertRow)]
    headers   = [c.replace("_", " ").title() for c in col_names]

    hdr_font  = Font(name="Calibri", bold=True, color=HDR_FG)
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center", indent=1)

    # Header row
    ws.row_dimensions[1].height = 32
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align

    # Data rows
    for ri, row in enumerate(rows, 2):
        row_dict = asdict(row)
        # Alternate row background
        row_bg = PatternFill("solid", fgColor=GREY_BG) if ri % 2 == 0 else None
        for ci, col in enumerate(col_names, 1):
            val = row_dict[col]
            # Convert days_until_expiry to int for numeric sort in Excel
            if col == "days_until_expiry" and val:
                try:
                    val = int(val)
                except ValueError:
                    pass
            if col == "port" and val:
                try:
                    val = int(val)
                except ValueError:
                    pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = center if col in (
                "port", "scan_status", "hostname_match", "wildcard",
                "self_signed", "chain_trusted", "is_ca", "hsts",
                "http_to_https_redirect", "deprecated_tls10", "deprecated_tls11",
                "expiry_status", "days_until_expiry", "http_status",
            ) else left
            if row_bg:
                cell.fill = row_bg

    last_row = len(rows) + 1

    # Conditional formatting – tls_deprecated_enabled (non-empty = amber)
    tls_dep_col_idx = col_names.index("tls_deprecated_enabled") + 1
    tls_dep_col_ltr = get_column_letter(tls_dep_col_idx)
    ws.conditional_formatting.add(
        f"{tls_dep_col_ltr}2:{tls_dep_col_ltr}{last_row}",
        CellIsRule(operator="notEqual", formula=['""'],
                   fill=PatternFill("solid", fgColor=AMBER_BG))
    )

    # Conditional formatting – expiry_status column
    expiry_col_idx = col_names.index("expiry_status") + 1
    expiry_col_ltr = get_column_letter(expiry_col_idx)
    rng_expiry = f"{expiry_col_ltr}2:{expiry_col_ltr}{last_row}"

    ws.conditional_formatting.add(rng_expiry, CellIsRule(
        operator="equal", formula=['"EXPIRED"'],
        fill=PatternFill("solid", fgColor="FF4444"),
        font=Font(bold=True, color="FFFFFF"),
    ))
    ws.conditional_formatting.add(rng_expiry, CellIsRule(
        operator="equal", formula=['"WARNING"'],
        fill=PatternFill("solid", fgColor="FFAA00"),
        font=Font(bold=True, color="FFFFFF"),
    ))
    ws.conditional_formatting.add(rng_expiry, CellIsRule(
        operator="equal", formula=['"OK"'],
        fill=PatternFill("solid", fgColor="44BB44"),
        font=Font(bold=True, color="FFFFFF"),
    ))

    # Conditional formatting – issues column (non-empty = red tint)
    issues_col_idx = col_names.index("issues") + 1
    issues_col_ltr = get_column_letter(issues_col_idx)
    rng_issues = f"{issues_col_ltr}2:{issues_col_ltr}{last_row}"
    ws.conditional_formatting.add(rng_issues, CellIsRule(
        operator="notEqual", formula=['""'],
        fill=PatternFill("solid", fgColor=RED_BG),
    ))

    # Conditional formatting – weaknesses column (non-empty = amber tint)
    weak_col_idx = col_names.index("weaknesses") + 1
    weak_col_ltr = get_column_letter(weak_col_idx)
    rng_weak = f"{weak_col_ltr}2:{weak_col_ltr}{last_row}"
    ws.conditional_formatting.add(rng_weak, CellIsRule(
        operator="notEqual", formula=['""'],
        fill=PatternFill("solid", fgColor=AMBER_BG),
    ))

    # Column widths
    for ci, col in enumerate(col_names, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _col_width(rows, col, headers[ci-1])

    # Freeze header + Excel Table
    ws.freeze_panes = "A2"
    table_ref = f"A1:{get_column_letter(len(col_names))}{last_row}"
    tbl = Table(displayName="CertScan", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(tbl)

    # ── Sheet 2: Issues summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues & Warnings")
    ws2.row_dimensions[1].height = 28
    sum_cols = ["url", "host", "expiry_status", "days_until_expiry",
                "chain_trusted", "hostname_match", "self_signed",
                "deprecated_tls10", "deprecated_tls11",
                "hsts", "issues", "weaknesses"]
    sum_hdrs = [c.replace("_", " ").title() for c in sum_cols]

    for ci, h in enumerate(sum_hdrs, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align

    flagged = [r for r in rows
               if (asdict(r)["issues"] or asdict(r)["weaknesses"])
               and asdict(r)["scan_status"] == "ok"]
    flagged.sort(key=lambda r: (
        0 if asdict(r)["expiry_status"] == "EXPIRED" else
        1 if asdict(r)["expiry_status"] == "WARNING" else 2,
        int(asdict(r)["days_until_expiry"]) if str(asdict(r)["days_until_expiry"]).lstrip("-").isdigit() else 9999,
    ))

    for ri, row in enumerate(flagged, 2):
        rd = asdict(row)
        row_bg = PatternFill("solid", fgColor=GREY_BG) if ri % 2 == 0 else None
        for ci, col in enumerate(sum_cols, 1):
            val = rd[col]
            if col == "days_until_expiry" and val:
                try:
                    val = int(val)
                except ValueError:
                    pass
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = left
            if row_bg:
                cell.fill = row_bg

    last2 = len(flagged) + 1
    for ci, col in enumerate(sum_cols, 1):
        max_len = max(
            (len(str(asdict(r)[col] or "")) for r in flagged),
            default=0,
        )
        ws2.column_dimensions[get_column_letter(ci)].width = min(
            max(max_len, len(sum_hdrs[ci-1])) + 2, 55
        )

    ws2.freeze_panes = "A2"
    if flagged:
        tbl2 = Table(displayName="IssuesSummary",
                     ref=f"A1:{get_column_letter(len(sum_cols))}{last2}")
        tbl2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True,
        )
        ws2.add_table(tbl2)

    # ── Sheet 3: Stats ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Stats")
    total   = len(rows)
    expired = sum(1 for r in rows if asdict(r)["expiry_status"] == "EXPIRED")
    warning = sum(1 for r in rows if asdict(r)["expiry_status"] == "WARNING")
    ok_cert = sum(1 for r in rows if asdict(r)["expiry_status"] == "OK")
    no_ssl       = sum(1 for r in rows if asdict(r)["scan_status"] in ("no_ssl", "unreachable", "error"))
    unreachable  = sum(1 for r in rows if asdict(r)["scan_status"] == "unreachable")
    timed_out    = sum(1 for r in rows if "timeout" in (asdict(r)["error_detail"] or "").lower())
    trusted = sum(1 for r in rows if asdict(r)["chain_trusted"] == "yes")
    hsts_ok = sum(1 for r in rows if asdict(r)["hsts"] == "yes")
    redir   = sum(1 for r in rows if asdict(r)["http_to_https_redirect"] == "yes")
    dep10   = sum(1 for r in rows if asdict(r)["deprecated_tls10"] == "yes")
    dep11   = sum(1 for r in rows if asdict(r)["deprecated_tls11"] == "yes")

    stats = [
        ("Metric",                    "Count", "% of total"),
        ("Total URLs scanned",         total,   "100%"),
        ("Certificate OK",             ok_cert, f"{ok_cert/total*100:.0f}%" if total else "-"),
        ("Expiring (warning)",         warning, f"{warning/total*100:.0f}%" if total else "-"),
        ("Expired",                    expired, f"{expired/total*100:.0f}%" if total else "-"),
        ("No SSL / unreachable",       no_ssl,      f"{no_ssl/total*100:.0f}%" if total else "-"),
        ("  └─ unreachable (TCP)",      unreachable, f"{unreachable/total*100:.0f}%" if total else "-"),
        ("  └─ of which timed out",     timed_out,   f"{timed_out/total*100:.0f}%" if total else "-"),
        ("Chain trusted",              trusted, f"{trusted/total*100:.0f}%" if total else "-"),
        ("HSTS enabled",               hsts_ok, f"{hsts_ok/total*100:.0f}%" if total else "-"),
        ("HTTP→HTTPS redirect",        redir,   f"{redir/total*100:.0f}%" if total else "-"),
        ("TLS 1.0 still accepted",     dep10,   f"{dep10/total*100:.0f}%" if total else "-"),
        ("TLS 1.1 still accepted",     dep11,   f"{dep11/total*100:.0f}%" if total else "-"),
    ]

    ws3.row_dimensions[1].height = 28
    for ri, (label, count, pct) in enumerate(stats, 1):
        ws3.cell(ri, 1, label).font  = Font(name="Calibri", bold=(ri == 1), size=11)
        ws3.cell(ri, 2, count).font  = Font(name="Calibri", bold=(ri == 1), size=11)
        ws3.cell(ri, 3, pct).font    = Font(name="Calibri", bold=(ri == 1), size=11)
        if ri == 1:
            for ci in range(1, 4):
                ws3.cell(ri, ci).fill      = hdr_fill
                ws3.cell(ri, ci).font      = hdr_font
                ws3.cell(ri, ci).alignment = hdr_align
        ws3.cell(ri, 2).alignment = Alignment(horizontal="center")
        ws3.cell(ri, 3).alignment = Alignment(horizontal="center")

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12


    # ── Sheet 4: Unreachable / No-SSL hosts ───────────────────────────────────
    ws4 = wb.create_sheet("Unreachable & No-SSL")
    ws4.row_dimensions[1].height = 28
    skip_cols = ["url", "host", "port", "scan_status", "error_detail", "scanned_at"]
    skip_hdrs = [c.replace("_", " ").title() for c in skip_cols]
    for ci, h in enumerate(skip_hdrs, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
    skipped = [r for r in rows if asdict(r)["scan_status"] != "ok"]
    skipped.sort(key=lambda r: (asdict(r)["scan_status"], asdict(r)["host"]))
    for ri, row in enumerate(skipped, 2):
        rd = asdict(row)
        for ci, col in enumerate(skip_cols, 1):
            cell = ws4.cell(row=ri, column=ci, value=rd[col])
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = left
    for ci, col in enumerate(skip_cols, 1):
        max_len = max((len(str(asdict(r)[col] or "")) for r in skipped), default=0)
        ws4.column_dimensions[get_column_letter(ci)].width = min(max(max_len, len(skip_hdrs[ci-1])) + 2, 60)
    ws4.freeze_panes = "A2"
    if skipped:
        tbl4 = Table(displayName="UnreachableHosts",
                     ref=f"A1:{get_column_letter(len(skip_cols))}{len(skipped)+1}")
        tbl4.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws4.add_table(tbl4)

    wb.save(path)
    log.info("XLSX saved → %s  (%d rows, %d flagged)", path, total, len(flagged))


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="SSL certificate monitor — reads URL list, outputs CSV + XLSX.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("-i", "--input",     required=True,
                   help="Input TXT file (discovered_urls.txt or any URL list)")
    p.add_argument("-o", "--output",    default="ssl_report",
                   help="Output base name (extensions .xlsx and .csv added automatically)")
    p.add_argument("--warn-days",       type=int, default=DEFAULT_WARNING_DAYS,
                   help="Days-before-expiry threshold for WARNING status")
    p.add_argument("--threads",         type=int, default=DEFAULT_THREADS)
    p.add_argument("--timeout",         type=int, default=DEFAULT_TIMEOUT)
    p.add_argument("--csv-only",        action="store_true",
                   help="Skip XLSX generation (faster, no openpyxl needed)")
    p.add_argument("-v", "--verbose",   action="store_true")
    return p


def main() -> None:
    args = build_parser().parse_args()
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    input_path = Path(args.input)
    if not input_path.exists():
        log.error("Input file not found: %s", input_path)
        sys.exit(1)

    entries = load_urls(input_path)
    if not entries:
        log.error("No valid URLs found in %s", input_path)
        sys.exit(1)

    log.info("Loaded %d unique hosts to scan", len(entries))

    results: List[CertRow] = []
    with cf.ThreadPoolExecutor(max_workers=args.threads) as ex:
        futs = {
            ex.submit(scan_host, url, host, port, args.warn_days, args.timeout): (url, host)
            for url, host, port in entries
        }
        done = 0
        total = len(futs)
        for fut in cf.as_completed(futs):
            done += 1
            url, host = futs[fut]
            try:
                row = fut.result()
                results.append(row)
                _lvl = (
                    logging.DEBUG if row.scan_status in ("unreachable", "no_ssl")
                    else logging.INFO
                )
                log.log(
                    _lvl,
                    "[%d/%d] %-45s  status=%-12s  expiry=%-12s  days=%-5s  issues=%s",
                    done, total, host,
                    row.scan_status,
                    row.expiry_status, row.days_until_expiry,
                    row.issues or "none",
                )
            except Exception as e:
                log.error("Unexpected error scanning %s: %s", host, e)
                results.append(CertRow(
                    url=url, host=host, scan_status="error",
                    error_detail=str(e), scanned_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                ))

    results.sort(key=lambda r: (
        0 if r.expiry_status == "EXPIRED"  else
        1 if r.expiry_status == "WARNING"  else
        2 if r.expiry_status == ""         else 3,
        int(r.days_until_expiry) if r.days_until_expiry.lstrip("-").isdigit() else 9999,
        r.host,
    ))

    base = Path(args.output)
    csv_path  = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")

    write_csv(results, csv_path)
    if not args.csv_only:
        write_xlsx(results, xlsx_path, args.warn_days)

    # Summary to stdout
    expired = sum(1 for r in results if r.expiry_status == "EXPIRED")
    warning = sum(1 for r in results if r.expiry_status == "WARNING")
    print(json.dumps({
        "total_scanned": len(results),
        "expired":       expired,
        "warning":       warning,
        "ok":            sum(1 for r in results if r.expiry_status == "OK"),
        "no_ssl_errors": sum(1 for r in results if r.scan_status != "ok"),
        "csv":           str(csv_path),
        "xlsx":          str(xlsx_path) if not args.csv_only else "skipped",
    }, indent=2))


if __name__ == "__main__":
    main()
